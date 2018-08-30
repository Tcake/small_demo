#!/usr/bin/python
# encoding:utf-8

import openpyxl
import random
import requests
import json
import urllib
from copy import deepcopy
from time import sleep


class Rinse(object):
    def __init__(self, file_name, MAX_COL, MAX_ROW, COL_NAME, COL_NUM):
        self.FILENAME = file_name
        self.URL = 'http://dev01.haizhi.com:14396'
        self.TOKEN = '8085672ca918435172a1437c52bc4aa7'
        self.INIT_FIELD = {"remark": "",  # 表格结构数据 同户人
                      "name": self.FILENAME,
                      "ds_id": 'ds_54a6438f92c04e548a61bf5c242a44d2',
                      "schema": []}
        self.INT_A_COL = {"remark": "", "type": "string", "name": "", "title": ""}
        self.TB_ID = ''
        self.FIELD = []

        FILE = '/Users/zlt/small_demo/rinse/data/' + self.FILENAME + '.xlsx'
        self.MAX_COL = MAX_COL
        self.MAX_ROW = MAX_ROW
        self.COL_NAME = COL_NAME
        self.COL_NUM = COL_NUM

        file_wr = openpyxl.load_workbook(FILE, read_only=True)  # 打开xls文件

        self.table = file_wr.active

    def num(self):
        result = ''
        example = u'12344567890X'
        for i in range(18):
            if i == 17:
                result += random.choice(example)
            else:
                result += random.choice(example[:-1])
        return result

    def name(self):
        result = ''
        example = u'王庆国孟广好高晓军刘宝刚张小芬张金清李瑞欣孙荣保唐国珍金龙宇晶'
        for i in range(3):
            result += random.choice(example)
        return result

    def init_bdp(self):
        for first in self.table.iter_rows(max_row=1, max_col=self.MAX_COL):
            for cell in first:
                self.FIELD.append(cell.value)
                a_field = deepcopy(self.INT_A_COL)
                a_field.update({'name': cell.value})
                self.INIT_FIELD['schema'].append(a_field)
        self.TB_ID = self.create_tb(self.INIT_FIELD)

    def create_tb(self, data):
        url = "{}/api/tb/create?access_token={}".format(self.URL, self.TOKEN)
        result = json.loads(requests.post(url, json.dumps(data)).content)
        if not result['status'] == '0':
            print('创建工作表失败:{}'.format(result['errstr']))
            raise ValueError
        tb_id = str(result['result']['tb_id'])
        self.commit(tb_id)
        return tb_id

    def insert_data(self, data):
        fiels_quote = urllib.quote(json.dumps(self.FIELD))
        url = "{}/api/data/insert?access_token={}&tb_id={}&fields={}".format(self.URL, self.TOKEN, self.TB_ID, fiels_quote)

        flag = 0
        while flag < 3:
            try:
                result = json.loads(requests.post(url, json.dumps(data)).content)
            except requests.ConnectionError as e:
                print('连接失败了，我休息一会。{}'.format(e))
                flag += 1
                sleep(3)
            else:
                break

        if not result['status'] == '0':
            print('插入数据失败:{}'.format(result['errstr']))
            raise ValueError

    def create_ds(self, name='公安业务demo数据'):
        global DS_ID
        url = "{}/api/ds/create?access_token={}&name={}".format(self.URL, self.TOKEN, name)
        result = json.loads(requests.get(url).content)
        if not result['status'] == '0':
            print('创建数据库失败:{}'.format(result['errstr']))
            raise ValueError
        DS_ID = str(result['result']['ds_id'])
        print('创建成功，数据库id:'+DS_ID)

    def commit(self, tb_id):
        url = "{}/api/tb/commit?access_token={}&tb_id={}".format(self.URL, self.TOKEN, tb_id)

        flag = 0
        while flag < 3:
            try:
                result = json.loads(requests.get(url).content)
            except requests.ConnectionError as e:
                print('连接失败了，我休息一会。{}'.format(e))
                flag += 1
                sleep(3)
            else:
                break

        if not result['status'] == '0':
                print('commit数据失败:{}'.format(result['errstr']))
                raise ValueError

    def main(self):
        print ('{}文件打开完成，开始任务'.format(self.FILENAME))
        self.init_bdp()
        print('bdp表建立完成，表id：{}，开始插入数据'.format(self.TB_ID))
        item = []
        count = 0
        for row in self.table.iter_rows(min_row=2, max_row=self.MAX_ROW, max_col=self.MAX_COL):
            arow = []
            for cell in row:
                arow.append(cell.value)

            a_name = self.name()
            for acol in self.COL_NAME:
                arow[acol - 1] = a_name
            a_num = self.num()
            for acol in self.COL_NUM:
                arow[acol - 1] = a_num

            item.append(arow)
            count += 1
            if count % 20000 == 0:
                self.insert_data(item)
                print('数据行数：{}，插入数据成功！'.format(count))
                item = []
        if item:
            self.insert_data(item)
        print('正在commit')
        self.commit(self.TB_ID)
        print('任务完成')


if __name__ == '__main__':
    a = Rinse('DW_二手车交易信息',2,3,4,5)
    a.commit('tb_b2c3ebb0a1184e1eb676676cae066f71')

